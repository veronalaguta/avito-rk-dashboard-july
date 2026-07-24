#!/usr/bin/env python3
import hashlib
import json
import re
from collections import defaultdict
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl
import requests
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
AVITO_PRO_PATH = Path("/Users/veronikalagutkina/Downloads/Орлов Статистика_с_2026-07-01_по_2026-07-23.xlsx")
XML_SHEET_PATH = Path("/Users/veronikalagutkina/Downloads/Холодильник СПБ Николай Орлов.xlsx")
PREVIEW_DIR = ROOT / "assets" / "orlov-all-previews"
DATA_DIR = ROOT / "data" / "processed"

SHEETS = [
    ("Avito Кондиционеры", "Кондиционеры"),
    ("Avito Электрика", "Электрика"),
    ("Avito Сантехника", "Сантехника"),
    ("Avito Холодильники", "Холодильники"),
]


def clean_header(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def plain_html(value):
    text = clean_text(value)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"</p\s*>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_hyperlink(value):
    text = clean_text(value)
    match = re.match(r'=HYPERLINK\("([^"]+)","([^"]+)"\)', text)
    if match:
        return match.group(2), match.group(1)
    return text, ""


def numeric(value):
    if value in ("", None):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    return float(text) if text else 0.0


def first_url(value):
    text = clean_text(value)
    if not text:
        return ""
    for line in re.split(r"[\n\r]+", text):
        line = line.strip()
        if line:
            if line.startswith("http://avito.ru/"):
                return line.replace("http://", "https://", 1)
            return line
    return ""


def normalize_image_url(url):
    if not url:
        return ""
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    slug = query.get("imageSlug", [""])[0]
    if "autoload/1/items-to-feed/images" in parsed.path and slug:
        return "https://00.img.avito.st" + slug
    if url.startswith("http://avito.ru/"):
        return url.replace("http://", "https://", 1)
    return url


def parse_avito_pro():
    wb = openpyxl.load_workbook(AVITO_PRO_PATH, data_only=False, read_only=True)
    ws = wb.active
    ws.reset_dimensions()
    rows = ws.iter_rows(values_only=True)
    headers = [clean_header(x) for x in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}
    parsed = {}
    for row in rows:
        ad_id, url = parse_hyperlink(row[idx["Номер объявления"]])
        if not ad_id:
            continue
        title, title_url = parse_hyperlink(row[idx["Название объявления"]])
        parsed[ad_id] = {
            "id": ad_id,
            "url": url or title_url,
            "region": clean_text(row[idx["Регион размещения"]]),
            "city": clean_text(row[idx["Город"]]),
            "address": clean_text(row[idx["Адрес"]]),
            "category": clean_text(row[idx["Категория"]]),
            "subcategory": clean_text(row[idx["Подкатегория"]]),
            "direction": clean_text(row[idx["Параметр"]]),
            "title": title,
            "price": clean_text(row[idx["Цена"]]),
            "first_publication_date": clean_text(row[idx["Дата первой публикации"]]),
            "days_on_avito": numeric(row[idx["Дней на Авито"]]),
            "shows": numeric(row[idx["Показы"]]),
            "ctr": numeric(row[idx["Конверсия из показов в просмотры"]]),
            "views": numeric(row[idx["Просмотры"]]),
            "view_price": numeric(row[idx["Средняя цена просмотра"]]),
            "contact_rate": numeric(row[idx["Конверсия из просмотров в контакты"]]),
            "target_views": numeric(row[idx["Целевые просмотры"]]),
            "contacts": numeric(row[idx["Контакты"]]),
            "chat": numeric(row[idx["Написали в чат"]]),
            "phone": numeric(row[idx["Посмотрели телефон"]]),
            "contact_price": numeric(row[idx["Средняя цена контакта"]]),
            "favorites": numeric(row[idx["Добавили в избранное"]]),
            "spend": numeric(row[idx["Расходы на объявления"]]),
        }
    return parsed


def parse_xml_sheet():
    wb = openpyxl.load_workbook(XML_SHEET_PATH, data_only=True, read_only=True)
    parsed = {}
    for sheet_name, direction in SHEETS:
        ws = wb[sheet_name]
        ws.reset_dimensions()
        rows = ws.iter_rows(values_only=True)
        headers = [clean_header(x) for x in next(rows)]
        idx = {h: i for i, h in enumerate(headers) if h}
        required = ["Номер объявления", "Заголовок объявления", "Цена", "Ссылки на фото"]
        if not all(x in idx for x in required):
            raise RuntimeError(f"Missing required columns in {sheet_name}: {required}")
        for row in rows:
            ad_id = clean_text(row[idx["Номер объявления"]]) if idx["Номер объявления"] < len(row) else ""
            if not ad_id or not re.fullmatch(r"\d{6,}", ad_id):
                continue
            images = clean_text(row[idx["Ссылки на фото"]])
            parsed[ad_id] = {
                "id": ad_id,
                "xml_sheet": sheet_name,
                "xml_direction": direction,
                "xml_url": clean_text(row[idx.get("Ссылка на объявление", -1)]) if idx.get("Ссылка на объявление", -1) >= 0 else "",
                "xml_status": clean_text(row[idx.get("Статус на Avito", -1)]) if idx.get("Статус на Avito", -1) >= 0 else "",
                "xml_title": clean_text(row[idx["Заголовок объявления"]]),
                "xml_text": plain_html(row[idx.get("Описание", -1)]) if idx.get("Описание", -1) >= 0 else "",
                "xml_price": clean_text(row[idx["Цена"]]),
                "xml_images": images,
                "first_image_url": first_url(images),
                "image_folder": clean_text(row[idx.get("Название папки с фотографиями", -1)]) if idx.get("Название папки с фотографиями", -1) >= 0 else "",
                "address_xml": clean_text(row[idx.get("Адрес", -1)]) if idx.get("Адрес", -1) >= 0 else "",
                "metro_xml": clean_text(row[idx.get("Улица Район Метро", -1)]) if idx.get("Улица Район Метро", -1) >= 0 else "",
            }
    return parsed


def ext_from_response(url, content_type):
    parsed = urlparse(url)
    suffix = Path(parsed.path).suffix.lower()
    if suffix in [".jpg", ".jpeg", ".png", ".webp"]:
        return ".jpg" if suffix == ".jpeg" else suffix
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    return ".jpg"


def image_hash(path):
    with Image.open(path) as img:
        img = img.convert("L").resize((8, 8))
        pixels = list(img.getdata())
    avg = sum(pixels) / len(pixels)
    bits = "".join("1" if p >= avg else "0" for p in pixels)
    return f"{int(bits, 2):016x}"


def download_preview(url):
    url = normalize_image_url(url)
    if not url:
        return "", "", ""
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    url_digest = hashlib.sha256(url.encode("utf-8")).hexdigest()[:16]
    cached = sorted(PREVIEW_DIR.glob(f"{url_digest}-*"))
    if cached:
        path = cached[0]
        file_digest = path.stem.split("-", 1)[1]
        return str(path.relative_to(ROOT)), file_digest, image_hash(path)
    response = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"}, stream=True)
    response.raise_for_status()
    content = response.content
    file_digest = hashlib.sha256(content).hexdigest()[:16]
    ext = ext_from_response(response.url, response.headers.get("content-type", ""))
    path = PREVIEW_DIR / f"{url_digest}-{file_digest}{ext}"
    if not path.exists():
        path.write_bytes(content)
    return str(path.relative_to(ROOT)), file_digest, image_hash(path)


def classify_group(group):
    ctr = group["views"] / group["shows"] if group["shows"] else 0
    cr = group["contacts"] / group["views"] if group["views"] else 0
    if ctr >= 0.12 and cr >= 0.25:
        return "scale", "сильный вход и хорошая контактность"
    if ctr >= 0.12 and cr < 0.12:
        return "fix_offer", "превью цепляет, но связка после просмотра слабая"
    if ctr < 0.07 and cr >= 0.25:
        return "fix_preview", "объявление убеждает, но превью недобирает вход"
    if ctr < 0.07 and cr < 0.12:
        return "stop_or_rework", "слабый вход и слабая контактность"
    return "watch", "средняя зона: проверять по гео и объему"


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    pro = parse_avito_pro()
    xml = parse_xml_sheet()
    ads = []
    for ad_id, pro_row in pro.items():
        xml_row = xml.get(ad_id, {})
        merged = {**pro_row, **xml_row}
        image_url = merged.get("first_image_url", "")
        try:
            local_image, file_hash, visual_hash = download_preview(image_url)
        except Exception as exc:
            local_image, file_hash, visual_hash = "", "", ""
            merged["image_error"] = str(exc)
        merged["local_image"] = local_image
        merged["image_file_hash"] = file_hash
        merged["image_visual_hash"] = visual_hash
        merged["title"] = merged.get("xml_title") or merged.get("title") or ""
        merged["text"] = merged.get("xml_text") or ""
        merged["text_hash"] = hashlib.sha256(merged["text"].encode("utf-8")).hexdigest()[:12] if merged["text"] else ""
        merged["text_excerpt"] = merged["text"][:220]
        merged["price"] = merged.get("xml_price") or merged.get("price") or ""
        merged["direction"] = merged.get("xml_direction") or merged.get("direction") or ""
        ads.append(merged)

    groups = {}
    for ad in ads:
        key = ad.get("image_file_hash") or f"missing:{ad['id']}"
        if key not in groups:
            groups[key] = {
                "preview_key": key,
                "image": ad.get("local_image", ""),
                "first_image_url": ad.get("first_image_url", ""),
                "ads_count": 0,
                "shows": 0.0,
                "views": 0.0,
                "contacts": 0.0,
                "spend": 0.0,
                "favorites": 0.0,
                "chat": 0.0,
                "phone": 0.0,
                "directions": set(),
                "cities": set(),
                "titles": defaultdict(int),
                "folders": defaultdict(int),
                "ads": [],
            }
        g = groups[key]
        g["ads_count"] += 1
        for metric in ["shows", "views", "contacts", "spend", "favorites", "chat", "phone"]:
            g[metric] += float(ad.get(metric) or 0)
        g["directions"].add(ad.get("direction", ""))
        g["cities"].add(ad.get("city", ""))
        g["titles"][ad.get("title", "")] += 1
        g["folders"][ad.get("image_folder", "")] += 1
        g["ads"].append({
            "id": ad["id"],
            "url": ad.get("url") or ad.get("xml_url", ""),
            "title": ad.get("title", ""),
            "direction": ad.get("direction", ""),
            "city": ad.get("city", ""),
            "shows": ad.get("shows", 0),
            "views": ad.get("views", 0),
            "contacts": ad.get("contacts", 0),
            "spend": ad.get("spend", 0),
            "image_folder": ad.get("image_folder", ""),
        })

    output_groups = []
    for g in groups.values():
        ctr = g["views"] / g["shows"] if g["shows"] else 0
        contact_rate = g["contacts"] / g["views"] if g["views"] else 0
        contact_price = g["spend"] / g["contacts"] if g["contacts"] else None
        status, status_text = classify_group(g)
        top_title = max(g["titles"].items(), key=lambda x: x[1])[0] if g["titles"] else ""
        top_folder = max(g["folders"].items(), key=lambda x: x[1])[0] if g["folders"] else ""
        g["ads"].sort(key=lambda x: (x["contacts"], x["views"], x["shows"]), reverse=True)
        output_groups.append({
            "preview_key": g["preview_key"],
            "image": g["image"],
            "first_image_url": g["first_image_url"],
            "ads_count": g["ads_count"],
            "shows": round(g["shows"], 2),
            "views": round(g["views"], 2),
            "contacts": round(g["contacts"], 2),
            "spend": round(g["spend"], 2),
            "favorites": round(g["favorites"], 2),
            "chat": round(g["chat"], 2),
            "phone": round(g["phone"], 2),
            "ctr": ctr,
            "contact_rate": contact_rate,
            "contact_price": contact_price,
            "directions": sorted(x for x in g["directions"] if x),
            "cities": sorted(x for x in g["cities"] if x),
            "top_title": top_title,
            "top_folder": top_folder,
            "status": status,
            "status_text": status_text,
            "ads": g["ads"][:8],
        })

    combo_groups = {}
    for ad in ads:
        if not ad.get("image_file_hash"):
            continue
        key = "|".join([
            ad.get("city", ""),
            ad.get("image_file_hash", ""),
            ad.get("title", ""),
            ad.get("text_hash", ""),
        ])
        if key not in combo_groups:
            combo_groups[key] = {
                "combo_key": key,
                "city": ad.get("city", ""),
                "direction": ad.get("direction", ""),
                "image": ad.get("local_image", ""),
                "title": ad.get("title", ""),
                "text_hash": ad.get("text_hash", ""),
                "text_excerpt": ad.get("text_excerpt", ""),
                "ads_count": 0,
                "shows": 0.0,
                "views": 0.0,
                "contacts": 0.0,
                "spend": 0.0,
                "ads": [],
            }
        cg = combo_groups[key]
        cg["ads_count"] += 1
        for metric in ["shows", "views", "contacts", "spend"]:
            cg[metric] += float(ad.get(metric) or 0)
        cg["ads"].append({
            "id": ad["id"],
            "url": ad.get("url") or ad.get("xml_url", ""),
            "shows": ad.get("shows", 0),
            "views": ad.get("views", 0),
            "contacts": ad.get("contacts", 0),
            "spend": ad.get("spend", 0),
        })

    output_combos = []
    for cg in combo_groups.values():
        ctr = cg["views"] / cg["shows"] if cg["shows"] else 0
        contact_rate = cg["contacts"] / cg["views"] if cg["views"] else 0
        contact_price = cg["spend"] / cg["contacts"] if cg["contacts"] else None
        output_combos.append({
            **cg,
            "shows": round(cg["shows"], 2),
            "views": round(cg["views"], 2),
            "contacts": round(cg["contacts"], 2),
            "spend": round(cg["spend"], 2),
            "ctr": ctr,
            "contact_rate": contact_rate,
            "contact_price": contact_price,
        })
    output_combos.sort(key=lambda x: (x["contacts"], x["views"], x["shows"]), reverse=True)

    output_groups.sort(key=lambda x: (x["contacts"], x["views"], x["shows"]), reverse=True)
    ads.sort(key=lambda x: (x.get("contacts", 0), x.get("views", 0), x.get("shows", 0)), reverse=True)
    public_ads = [
        {
            "id": ad["id"],
            "url": ad.get("url") or ad.get("xml_url", ""),
            "direction": ad.get("direction", ""),
            "category": ad.get("category", ""),
            "subcategory": ad.get("subcategory", ""),
            "city": ad.get("city", ""),
            "title": ad.get("title", ""),
            "first_publication_date": ad.get("first_publication_date", ""),
            "days_on_avito": ad.get("days_on_avito", 0),
            "text_hash": ad.get("text_hash", ""),
            "text_excerpt": ad.get("text_excerpt", ""),
            "image": ad.get("local_image", ""),
            "shows": ad.get("shows", 0),
            "ctr": ad.get("ctr", 0),
            "views": ad.get("views", 0),
            "view_price": ad.get("view_price", 0),
            "contact_rate": ad.get("contact_rate", 0),
            "contacts": ad.get("contacts", 0),
            "contact_price": ad.get("contact_price", 0),
            "spend": ad.get("spend", 0),
        }
        for ad in ads
    ]
    summary = {
        "account": "Николай Орлов ИП",
        "period": "2026-07-01 — 2026-07-23",
        "source_avito_pro": str(AVITO_PRO_PATH),
        "source_xml_sheet": str(XML_SHEET_PATH),
        "ads_total": len(ads),
        "ads_matched_xml": sum(1 for ad in ads if ad.get("first_image_url")),
        "preview_groups_total": len(output_groups),
        "preview_groups_with_image": sum(1 for g in output_groups if g.get("image")),
        "combo_groups_total": len(output_combos),
        "shows": round(sum(ad.get("shows", 0) for ad in ads), 2),
        "views": round(sum(ad.get("views", 0) for ad in ads), 2),
        "contacts": round(sum(ad.get("contacts", 0) for ad in ads), 2),
        "spend": round(sum(ad.get("spend", 0) for ad in ads), 2),
    }
    payload = {"summary": summary, "groups": output_groups, "combos": output_combos, "ads": public_ads}
    (DATA_DIR / "orlov_creatives.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "orlov_creatives.js").write_text(
        "window.ORLOV_CREATIVE_DATA = " + json.dumps(payload, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
